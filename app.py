from flask import Flask, render_template, request, send_file, session, redirect, url_for
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import date
import os

app = Flask(__name__)
app.secret_key = "clave_secreta_para_session"

PRESUPUESTOS = {
    "estacado": {
        "plantilla": "plantillas/estacado.xlsx",
        "campos": {
            "Varillas roscadas mm" : "B22",
            "Pulgadas Puntales ": "B26",
            "Estacado Metros lineales" : "B11",
            "Estacado Precio unitario": "D11",
            "Escalones cantidad" : "B13",
            "Escalones Precio Unitario": "D13"
        },
    },
    "nivelacion": {
        "plantilla": "plantillas/nivelacion.xlsx",
        "campos": {
            "Cantidad Camiones tierra negra con mano de obra": "B12",
            "Precio Camiones tierra negra con mano de obra": "D12",
            "Cantidad Camiones de Relleno": "B13",
            "Precio Camiones de Relleno": "D13",
            "Cantidad Mano de obra fina de tierra negra": "B14",
            "Precio Mano de obra fina de tierra negra": "D14",
            "Tipo de Pasto": "C16",
            "Cantidad Pasto": "B16",
            "Precio Pasto m2": "D16",
            "Precio Colocacion m2": "D17",
        },
    },
    "riego": {
        "plantilla": "plantillas/riego.xlsx",
        "campos": {
            "Costo Materiales de riego": "D10",
            "Precio Automatizacion": "D11",
            "Precio Mano de obra": "D12",
        },
    },
    "estructuras": {
        "plantilla": "plantillas/estructuras.xlsx"
    },
}

# 🔹 Función auxiliar para escribir valores en Excel
def completar_planilla(plantilla, data):
    wb = load_workbook(plantilla)
    ws = wb.active
    for campo, celda in data.items():
        ws[celda] = data[campo]
    return wb, ws


def generar_estacado(data):
    conf = PRESUPUESTOS["estacado"]

    wb = load_workbook(conf["plantilla"])
    ws = wb.active

    ubicacion = session.get("ubicacion", "")
    ws["A3"] = ubicacion
    ws["A5"] = date.today().strftime("%d/%m/%Y")

    # --- Cargar datos ---
    for campo, celda in conf["campos"].items():
        valor = data.get(campo)

        if not valor:
            continue

        try:
            #convertir a número si corresponde
            valor_num = float(valor)
            ws[celda] = valor_num

           
            if celda in ("D11", "D13"):
                ws[celda].number_format = '"$"#,##0.00'
        except ValueError:
            # texto
            ws[celda] = valor

    estacado_metros = float(data.get("Estacado Metros lineales") or 0)
    estacado_precio = float(data.get("Estacado Precio Unitario") or 0)
    escalones_cantidad = float(data.get("Escalones cantidad") or 0)
    escalones_precio = float(data.get("Escalones Precio Unitario") or 0)

    total = (
        estacado_metros * estacado_precio
        + escalones_cantidad * escalones_precio
    )

    nombre_archivo = f"estacado_{ubicacion}.xlsx"
    wb.save(nombre_archivo)

    return nombre_archivo, total


def generar_nivelacion(data):
    conf = PRESUPUESTOS["nivelacion"]
    wb = load_workbook(conf["plantilla"])
    ws = wb.active
    ubicacion = session.get("ubicacion", "")
    ws["A3"] = ubicacion

    for campo, celda in conf["campos"].items():
        if campo in data and data[campo]:
            ws[celda] = data[campo]
    ws["A5"] = date.today().strftime("%d/%m/%Y")

    try:
        total = (
            float(data.get("Cantidad Camiones tierra negra con mano de obra", 0)) * float(data.get("Precio Camiones tierra negra con mano de obra", 0))
            + float(data.get("Cantidad Camiones de Relleno", 0)) * float(data.get("Precio Camiones de Relleno", 0))
            + float(data.get("Cantidad Mano de obra fina de tierra negra", 0)) * float(data.get("Precio Mano de obra fina de tierra negra", 0))
            + float(data.get("Cantidad Pasto", 0)) * float(data.get("Precio Pasto")) 
            + float(data.get("Precio Colocacion", 0)) * float(data.get("Cantidad Pasto"))
        )
    except ValueError:
        total = 0

    nombre_archivo = f"nivelacion_{ubicacion}.xlsx"
    wb.save(nombre_archivo)
    return nombre_archivo, total


def generar_riego(data):
    conf = PRESUPUESTOS["riego"]
    wb = load_workbook(conf["plantilla"])
    ws = wb.active
    ubicacion = session.get("ubicacion", "")
    ws["A4"] = ubicacion

    for campo, celda in conf["campos"].items():
        if campo in data and data[campo]:
            ws[celda] = data[campo]
    ws["A5"] = date.today().strftime("%d/%m/%Y")

    try:
        total = (
            float(data.get("Costo Materiales de riego", 0))
            + float(data.get("Precio Automatizacion", 0))
            + float(data.get("Precio Mano de obra", 0))
        )
    except ValueError:
        total = 0

    nombre_archivo = f"riego_{ubicacion}.xlsx"
    wb.save(nombre_archivo)
    return nombre_archivo, total


# --- RUTAS FLASK ---

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        # 🔹 Guardar la ubicación una sola vez
        session["ubicacion"] = request.form["ubicacion"]
        return redirect(url_for("menu"))
    return render_template("index.html")


@app.route("/menu")
def menu():
    return render_template("menu.html", tipos=PRESUPUESTOS.keys(), ubicacion=session.get("ubicacion", ""))

@app.route("/estructuras", methods=["GET"])
def estructuras():
    return render_template("formulario_estructuras.html")  # tu HTML nuevo


@app.route("/formulario/<tipo>")
def formulario(tipo):
    if tipo not in PRESUPUESTOS:
        return "Tipo de presupuesto inválido", 400
    return render_template("formulario.html", tipo=tipo, campos=PRESUPUESTOS[tipo]["campos"].keys())


@app.route("/agregar", methods=["POST"])
def agregar():
    tipo = request.form["tipo"]
    data = dict(request.form)

    if "resumen" not in session:
        session["resumen"] = {}

    if tipo == "estacado":
        archivo, total = generar_estacado(data)
        session["resumen"]["total_estacado"] = total
    elif tipo == "nivelacion":
        archivo, total = generar_nivelacion(data)
        session["resumen"]["total_nivelacion"] = total
    elif tipo == "riego":
        archivo, total = generar_riego(data)
        session["resumen"]["total_riego"] = total

    session.modified = True
    return send_file(archivo, as_attachment=True)


@app.route("/resumen")
def resumen():
    resumen_data = session.get("resumen", {})
    ubicacion = session.get("ubicacion", "")
    wb = load_workbook("resumen.xlsx")
    ws = wb.active

    ws["A3"] = ubicacion
    ws["A5"] = date.today().strftime("%d/%m/%Y")
    ws["C11"] = resumen_data.get("total_nivelacion", 0)
    ws["C13"] = resumen_data.get("total_riego", 0)
    ws["C15"] = resumen_data.get("total_estacado", 0)

    resumen_filename = "resumen_general.xlsx"
    wb.save(resumen_filename)
    return send_file(resumen_filename, as_attachment=True)

@app.route("/agregar_estructuras", methods=["POST"])
def agregar_estructuras():
    conf = PRESUPUESTOS["estructuras"]
    ubicacion = session.get("ubicacion", "Sin ubicación")

    wb = load_workbook(conf["plantilla"])
    ws = wb.active

    ws["A3"] = f"Estructuras - {ubicacion}"
    ws["A5"] = date.today().strftime("%d/%m/%Y")

    estructuras = ["muelle", "descanso", "rampa", "pergola"]
    fila = 10
    total_general = 0

    for tipo in estructuras:
        largo = request.form.get(f"{tipo}_largo")
        ancho = request.form.get(f"{tipo}_ancho")
        precio = request.form.get(f"{tipo}_precio")
        cantidad = request.form.get(f"{tipo}_cantidad")

        # solo para el muelle, también toma las pulgadas y material
        if tipo == "muelle":
            columnas_a = request.form.get("pulgadas_columnas_a")
            columnas_b = request.form.get("pulgadas_columnas_b")
            clavaderas_a = request.form.get("pulgadas_clavaderas_a")
            clavaderas_b = request.form.get("pulgadas_clavaderas_b")
            soleras_a = request.form.get("pulgadas_soleras_a")
            soleras_b = request.form.get("pulgadas_soleras_b")
            material = request.form.get("material", "")

            detalle = (
                f"Muelle de {largo} x {ancho} m, "
                f"columnas de {columnas_a}x{columnas_b} pulg, "
                f"clavaderas de {clavaderas_a}x{clavaderas_b} pulg, "
                f"soleras de {soleras_a}x{soleras_b} pulg, "
                f"material: {material}"
            )
        else:
            detalle = f"{tipo.capitalize()} de {largo} x {ancho} m"

        # si no hay saltea
        if not (largo and ancho and precio and cantidad):
            continue

        largo = float(largo)
        ancho = float(ancho)
        precio = float(precio)
        cantidad = float(cantidad)
        total = precio * cantidad

        ws[f"C{fila}"] = detalle
        ws[f"D{fila}"] = precio
        ws[f"B{fila}"] = cantidad
        ws[f"H{fila}"] = total
        fila += 2  # espacio entre estructuras

        total_general += total

    nombre_archivo = f"estructuras_{ubicacion}.xlsx"
    wb.save(nombre_archivo)

    #guarda en sesión
    if "resumen" not in session:
        session["resumen"] = {}

    session["resumen"]["total_estructuras"] = total_general
    session.modified = True

    return send_file(nombre_archivo, as_attachment=True)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
